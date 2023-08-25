import os
import tkinter as tk
from tkinter import filedialog
from tkinter import font as tkfont
from tkinter import StringVar, OptionMenu
from openpyxl import load_workbook, Workbook


# Function to handle the file selection
def select_xlsx_file():
    global input_file, columns
    input_file = filedialog.askopenfilename(title="Select XLSX File", filetypes=[("XLSX Files", "*.xlsx")])
    if input_file:
        # Update the label to display the selected file
        selected_file_label.config(text=f'Selected File: {input_file}')
        # Populate the dropdown menu with column headers from the selected XLSX file
        columns = get_column_headers(input_file)
        update_column_dropdown()


# Function to get column headers from the selected XLSX file
def get_column_headers(file_path):
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    return headers


# Function to update the column dropdown menu
def update_column_dropdown():
    if columns:
        selected_column.set(columns[0])  # Set the default selection
        column_menu["menu"].delete(0, "end")  # Clear the current options
        for column in columns:
            column_menu["menu"].add_command(label=column, command=tk._setit(selected_column, column))


# Function to split the XLSX file based on the selected column
def split_xlsx():
    if input_file:
        selected_column_value = selected_column.get()
        if selected_column_value:
            workbook = load_workbook(input_file)
            sheet = workbook.active

            # Prompt the user to select a directory to save the split XLSX files
            output_directory = filedialog.askdirectory(title="Choose where you want to save the split XLSX files")
            if not output_directory:
                return  # User canceled the directory selection

            # Create separate XLSX files based on the selected column value
            output_files = {}
            for row in sheet.iter_rows(min_row=2):
                column_value = row[columns.index(selected_column_value)].value
                if column_value not in output_files:
                    output_files[column_value] = Workbook()
                    output_sheet = output_files[column_value].active
                    output_sheet.append(columns)  # Add headers
                output_sheet = output_files[column_value].active
                output_sheet.append([cell.value for cell in row])

            # Save all output files in the selected directory
            for value, wb in output_files.items():
                output_file_path = os.path.join(output_directory, f'{value}.xlsx')
                wb.save(output_file_path)

            # Close all output files
            for wb in output_files.values():
                wb.close()

            output_label.config(text=f'Success! XLSX files split based on "{selected_column_value}".')
            output_label.pack(pady=10)  # Show the label


# Create the main Tkinter window
root = tk.Tk()
root.title("XLSX Splitter")

# Set the window size
root.geometry("600x400")

# Center the window on the screen
window_width = root.winfo_reqwidth()
window_height = root.winfo_reqheight()
position_x = int((root.winfo_screenwidth() / 2) - (window_width / 2))
position_y = int((root.winfo_screenheight() / 2) - (window_height / 2))
root.geometry("+{}+{}".format(position_x, position_y))

# Create a custom font for labels
custom_font = tkfont.Font(family="Helvetica", size=16)

# Initialize variables
input_file = None
columns = []

# Label for the selected file
selected_file_label = tk.Label(root, text="Select an XLSX file:", font=custom_font)
selected_file_label.pack(pady=10)

# Button to select an XLSX file
select_file_button = tk.Button(root, text="Select XLSX File", command=select_xlsx_file, font=custom_font)
select_file_button.pack(pady=10)

# Dropdown menu for selecting the column
selected_column = StringVar(root)
column_menu = OptionMenu(root, selected_column, "No XLSX file selected")
column_menu.pack(pady=10)

# Button to split the XLSX file
split_button = tk.Button(root, text="Split XLSX", command=split_xlsx, font=custom_font)
split_button.pack(pady=10)

# Label for displaying the result
output_label = tk.Label(root, text="", font=custom_font)

# Start the Tkinter main loop
root.mainloop()
